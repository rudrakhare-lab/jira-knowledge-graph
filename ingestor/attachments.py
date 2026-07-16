"""Stage 1b — Jira attachment download + text extraction.

Downloads each attachment over the authenticated session and extracts plain
text based on MIME type. Extraction never crashes the ingestion run: any
failure is caught and recorded as ``extracted_text: null``.

System dependency:
    OCR for images requires the Tesseract binary:
        brew install tesseract        # macOS
        apt install tesseract-ocr     # Linux

Per-attachment output record (one per attachment in a ticket's ``attachments``):
    {
      "filename": "rca-report.pdf",
      "mime_type": "application/pdf",
      "size": 98304,
      "extracted_text": "Root cause: ...",   # or None
      "extraction_method": "pdf"              # pdf|ocr|docx|xlsx|text or None
    }
"""

from __future__ import annotations

import asyncio
import io
import logging
import os
from typing import Optional

import aiohttp

logger = logging.getLogger("ingestor.attachments")

# --- Limits / retry policy (mirrors fetch.py; kept here so this file is
#     independent of fetch.py per the agreed structure) -----------------------
MAX_ATTACHMENT_BYTES = 10_485_760  # 10 MB — larger files are skipped
RETRYABLE_STATUS = {429, 500, 502, 503, 504}
MAX_RETRIES = 3

# MIME types handled by each extractor.
PDF_TYPES = {"application/pdf"}
IMAGE_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/gif"}
DOCX_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
}
XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
}
TEXT_TYPES = {"text/plain", "text/csv", "text/x-log", "text/markdown"}
TEXT_EXTENSIONS = (".log", ".txt", ".md", ".csv")


# --- Text extractors (synchronous; run inline per the spec) -----------------
def _extract_pdf(data: bytes) -> str:
    from pdfminer.high_level import extract_text

    return extract_text(io.BytesIO(data)) or ""


def _extract_image_ocr(data: bytes) -> str:
    import pytesseract
    from PIL import Image

    with Image.open(io.BytesIO(data)) as img:
        return pytesseract.image_to_string(img) or ""


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append("\t".join(cells))
    finally:
        wb.close()
    return "\n".join(parts)


def _extract_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _select_method(mime_type: str, filename: str) -> Optional[str]:
    """Return the extraction method name for a (mime, filename), or None."""
    mime = (mime_type or "").lower()
    name = (filename or "").lower()
    if mime in PDF_TYPES:
        return "pdf"
    if mime in IMAGE_TYPES:
        return "ocr"
    if mime in DOCX_TYPES:
        return "docx"
    if mime in XLSX_TYPES:
        return "xlsx"
    if mime in TEXT_TYPES or name.endswith(TEXT_EXTENSIONS):
        return "text"
    return None


_EXTRACTORS = {
    "pdf": _extract_pdf,
    "ocr": _extract_image_ocr,
    "docx": _extract_docx,
    "xlsx": _extract_xlsx,
    "text": _extract_text,
}


async def _download(
    session: aiohttp.ClientSession, url: str
) -> Optional[bytes]:
    """Download attachment bytes over the authenticated session.

    Retries on 429 (honoring Retry-After), 5xx, and connection/timeout
    errors with exponential backoff. Returns ``None`` if all attempts fail.
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            async with session.get(url) as resp:
                if resp.status in RETRYABLE_STATUS:
                    delay = _retry_delay(resp, attempt)
                    logger.warning(
                        "attachment download %s -> HTTP %s, retry %d/%d in %.1fs",
                        url, resp.status, attempt, MAX_RETRIES, delay,
                    )
                    await asyncio.sleep(delay)
                    continue
                resp.raise_for_status()
                return await resp.read()
        except (aiohttp.ClientError, asyncio.TimeoutError) as exc:
            if attempt == MAX_RETRIES:
                logger.warning("attachment download failed %s: %s", url, exc)
                return None
            delay = 2 ** attempt
            logger.warning(
                "attachment download error %s: %s, retry %d/%d in %.1fs",
                url, exc, attempt, MAX_RETRIES, delay,
            )
            await asyncio.sleep(delay)
    return None


def _retry_delay(resp: aiohttp.ClientResponse, attempt: int) -> float:
    if resp.status == 429:
        retry_after = resp.headers.get("Retry-After")
        if retry_after:
            try:
                return float(retry_after)
            except ValueError:
                pass
    return float(2 ** attempt)


async def _process_one(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    attachment: dict,
) -> dict:
    """Download + extract a single attachment into an output record."""
    filename = attachment.get("filename", "")
    mime_type = attachment.get("mimeType", "")
    size = attachment.get("size", 0) or 0
    content_url = attachment.get("content")

    record = {
        "filename": filename,
        "mime_type": mime_type,
        "size": size,
        "extracted_text": None,
        "extraction_method": None,
    }

    if size > MAX_ATTACHMENT_BYTES:
        logger.info(
            "skip attachment %s (%d bytes > %d limit)",
            filename, size, MAX_ATTACHMENT_BYTES,
        )
        return record

    method = _select_method(mime_type, filename)
    if method is None:
        # Unsupported type (video, zip, exe, ...) — skip extraction.
        return record

    if not content_url:
        logger.warning("attachment %s has no content URL", filename)
        return record

    try:
        async with semaphore:
            data = await _download(session, content_url)
        if data is None:
            return record
        # Extraction is CPU-bound but brief; run inline per the spec.
        record["extracted_text"] = _EXTRACTORS[method](data)
        record["extraction_method"] = method
    except Exception as exc:  # never crash the run on a bad attachment
        logger.warning(
            "extraction failed for %s (%s): %s", filename, method, exc
        )
        record["extracted_text"] = None
        record["extraction_method"] = None
    return record


async def process_attachments(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    attachments_field: Optional[list],
) -> list[dict]:
    """Process a ticket's ``fields.attachment`` array into output records.

    Downloads are bounded by the shared ``semaphore`` (same one used for
    comments/changelog). Returns an empty list when there are no attachments.
    """
    if not attachments_field:
        return []
    tasks = [
        _process_one(session, semaphore, att) for att in attachments_field
    ]
    return await asyncio.gather(*tasks)
